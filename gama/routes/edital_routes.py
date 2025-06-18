from flask import Blueprint, render_template, request, redirect, url_for, flash, session
import openpyxl
from gama.models.edital import Edital, Cargo
from gama.controllers.edital_controller import EditalController
from gama.controllers.candidato_controller import CandidatoController

edital_bp = Blueprint('edital', __name__, template_folder='../templates')

@edital_bp.route('/gerenciar_editais')
def gerenciar_editais():
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para acessar esta página.', 'error')
        return redirect(url_for('auth.login'))

    editais = Edital.listar_todos()
    return render_template(
        'gerenciar_editais.html',
        editais=editais,
        nome=session.get('nome')
    )

@edital_bp.route('/adicionar_edital', methods=['POST'])
def adicionar_edital():
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para realizar esta operação.', 'error')
        return redirect(url_for('edital.gerenciar_editais'))

    numero_edital = request.form.get('numero_edital')
    data_edital = request.form.get('data_edital')
    data_publicacao = request.form.get('data_publicacao')
    vencimento_edital = request.form.get('vencimento_edital')
    cargos = request.form.get('cargos', '').split(',')

    if not numero_edital or not data_edital or not data_publicacao or not vencimento_edital:
        flash('Número, data do edital, data de publicação e vencimento são obrigatórios.', 'error')
        return redirect(url_for('edital.gerenciar_editais'))

    if not any(cargo.strip() for cargo in cargos):
        flash('Pelo menos um cargo é obrigatório.', 'error')
        return redirect(url_for('edital.gerenciar_editais'))

    edital_existente = EditalController.buscar_por_numero(numero_edital)
    if edital_existente:
        flash('Edital com este número já cadastrado.', 'error')
        return redirect(url_for('edital.gerenciar_editais'))

    sucesso = EditalController.criar_edital(numero_edital, data_edital, data_publicacao, vencimento_edital, cargos)

    if sucesso:
        flash('Edital adicionado com sucesso!', 'success')
    else:
        flash('Erro ao adicionar edital.', 'error')

    return redirect(url_for('edital.gerenciar_editais'))

@edital_bp.route('/editar_edital/<int:id_edital>', methods=['POST'])
def editar_edital(id_edital):
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para realizar esta operação.', 'error')
        return redirect(url_for('auth.login'))

    numero_edital = request.form.get('numero_edital')
    data_edital = request.form.get('data_edital')
    data_publicacao = request.form.get('data_publicacao')
    vencimento_edital = request.form.get('vencimento_edital')
    status = request.form.get('status')

    if not numero_edital:
        flash('Número do edital é obrigatório.', 'error')
        return redirect(url_for('edital.gerenciar_editais'))

    sucesso = EditalController.editar_edital(id_edital, numero_edital, data_edital, data_publicacao, vencimento_edital, status)

    if sucesso:
        flash('Edital atualizado com sucesso!', 'success')
    else:
        flash('Erro ao atualizar edital.', 'error')

    return redirect(url_for('edital.gerenciar_editais'))

@edital_bp.route('/prorrogar_edital/<int:id_edital>', methods=['POST'])
def prorrogar_edital(id_edital):
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para realizar esta operação.', 'error')
        return redirect(url_for('auth.login'))

    dias_prorrogacao = request.form.get('dias_prorrogacao')
    if not dias_prorrogacao or int(dias_prorrogacao) <= 0:
        flash('O prazo de prorrogação deve ser um número positivo.', 'error')
        return redirect(url_for('edital.gerenciar_editais'))

    sucesso = EditalController.prorrogar_edital(id_edital, dias_prorrogacao)

    if sucesso:
        flash('Edital prorrogado com sucesso!', 'success')
    else:
        flash('Erro ao prorrogar edital.', 'error')

    return redirect(url_for('edital.gerenciar_editais'))

@edital_bp.route('/excluir_edital/<int:id_edital>')
def excluir_edital(id_edital):
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para realizar esta operação.', 'error')
        return redirect(url_for('auth.login'))

    candidatos = CandidatoController.listar_por_edital(id_edital)
    if candidatos:
        flash('Não é possível excluir este edital pois possui candidatos cadastrados.', 'error')
        return redirect(url_for('edital.gerenciar_editais'))

    sucesso = EditalController.excluir_edital(id_edital)

    if sucesso:
        flash('Edital excluído com sucesso!', 'success')
    else:
        flash('Erro ao excluir edital.', 'error')

    return redirect(url_for('edital.gerenciar_editais'))

@edital_bp.route('/adicionar_candidato', methods=['POST'])
def adicionar_candidato():
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para realizar esta operação.', 'error')
        return redirect(url_for('auth.login'))

    nome = request.form.get('nome')
    numero_inscricao = request.form.get('numero_inscricao')
    id_cargo = request.form.get('id_cargo')
    ordem_nomeacao = request.form.get('ordem_nomeacao')
    pcd = request.form.get('pcd', '0')
    cotista = request.form.get('cotista', '0')
    id_edital = request.form.get('id_edital')
    nota = request.form.get('nota')

    pcd = 1 if pcd == '1' else 0
    cotista = 1 if cotista == '1' else 0

    try:
        nota = float(nota) if nota else None
        if nota is not None and nota < 0:
            flash('A nota não pode ser negativa.', 'error')
            return redirect(url_for('edital.gerenciar_editais'))
    except (ValueError, TypeError):
        flash('A nota deve ser um número válido.', 'error')
        return redirect(url_for('edital.gerenciar_editais'))

    if not nome or not numero_inscricao or not id_cargo or not ordem_nomeacao or not id_edital:
        flash('Nome, número de inscrição, cargo, ordem de nomeação e edital são obrigatórios.', 'error')
        return redirect(url_for('edital.gerenciar_editais'))

    candidato_existente = CandidatoController.obter_candidato(numero_inscricao)
    if candidato_existente:
        flash('Candidato com este número de inscrição já cadastrado.', 'error')
        return redirect(url_for('edital.gerenciar_editais'))

    sucesso = CandidatoController.criar_candidato(nome, numero_inscricao, id_cargo, ordem_nomeacao, pcd, cotista, id_edital, nota)

    if sucesso:
        flash('Candidato adicionado com sucesso!', 'success')
    else:
        flash('Erro ao adicionar candidato.', 'error')

    return redirect(url_for('edital.gerenciar_editais'))

@edital_bp.route('/editar_candidato/<int:id_candidato>', methods=['POST'])
def editar_candidato(id_candidato):
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para realizar esta operação.', 'error')
        return redirect(url_for('auth.login'))

    nome = request.form.get('nome')
    numero_inscricao = request.form.get('numero_inscricao')
    id_cargo = request.form.get('id_cargo')
    ordem_nomeacao = request.form.get('ordem_nomeacao')
    pcd = request.form.get('pcd', '0')
    cotista = request.form.get('cotista', '0')
    situacao = request.form.get('situacao')
    data_posse = request.form.get('data_posse') if situacao == 'nomeado' else None
    id_edital = request.form.get('id_edital')
    nota = request.form.get('nota')

    pcd = 1 if pcd == '1' else 0
    cotista = 1 if cotista == '1' else 0

    try:
        nota = float(nota) if nota else None
        if nota is not None and nota < 0:
            flash('A nota não pode ser negativa.', 'error')
            return redirect(url_for('edital.gerenciar_editais'))
    except (ValueError, TypeError):
        flash('A nota deve ser um número válido.', 'error')
        return redirect(url_for('edital.gerenciar_editais'))

    if not nome or not id_cargo or not ordem_nomeacao or not id_edital or not situacao:
        flash('Nome, cargo, ordem de nomeação, situação e edital são obrigatórios.', 'error')
        return redirect(url_for('edital.gerenciar_editais'))

    sucesso = CandidatoController.editar_candidato(id_candidato, nome, numero_inscricao, id_cargo, ordem_nomeacao, pcd, cotista, situacao, data_posse, id_edital, nota)

    if sucesso:
        flash('Candidato atualizado com sucesso!', 'success')
    else:
        flash('Erro ao atualizar candidato.', 'error')

    return redirect(url_for('edital.gerenciar_editais'))

@edital_bp.route('/excluir_candidato/<int:id_candidato>')
def excluir_candidato(id_candidato):
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para realizar esta operação.', 'error')
        return redirect(url_for('auth.login'))

    sucesso = CandidatoController.excluir_candidato(id_candidato)

    if sucesso:
        flash('Candidato excluído com sucesso!', 'success')
    else:
        flash('Erro ao excluir candidato.', 'error')

    return redirect(url_for('edital.gerenciar_editais'))

@edital_bp.route('/upload_candidatos', methods=['POST'])
def upload_candidatos():
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para realizar esta operação.', 'error')
        return redirect(url_for('auth.login'))

    id_edital = request.form.get('id_edital')
    arquivo = request.files.get('arquivo')

    if not id_edital or not arquivo:
        flash('Edital e arquivo são obrigatórios.', 'error')
        return redirect(url_for('edital.gerenciar_editais'))

    if not arquivo.filename.endswith('.xlsx'):
        flash('O arquivo deve ser no formato XLSX.', 'error')
        return redirect(url_for('edital.gerenciar_editais'))

    try:
        # Carregar o arquivo XLSX
        workbook = openpyxl.load_workbook(arquivo)
        sheet = workbook.active
        headers = [cell.value.strip().lower() if cell.value else '' for cell in next(sheet.iter_rows())]
        expected_columns = {'nome', 'numero_inscricao', 'ordem', 'pcd', 'cotista', 'cargo', 'nota'}

        if not expected_columns.issubset(headers):
            flash('O XLSX deve conter as colunas: nome, numero_inscricao, ordem, pcd, cotista, cargo, nota.', 'error')
            return redirect(url_for('edital.gerenciar_editais'))

        # Obter cargos disponíveis para o edital (normalizados para lowercase)
        cargos_disponiveis = {cargo[1].strip().lower(): cargo[0] for cargo in Cargo.listar_por_edital(id_edital)}
        
        candidatos = []
        row_num = 2  # Começa após o cabeçalho
        for row in sheet.iter_rows(min_row=2):
            row_data = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
            try:
                nome = str(row_data['nome']).strip() if row_data['nome'] else ''
                numero_inscricao = str(row_data['numero_inscricao']).strip() if row_data['numero_inscricao'] else ''
                ordem = int(row_data['ordem']) if row_data['ordem'] else 0
                pcd_value = str(row_data['pcd']).strip().lower() if row_data['pcd'] else '0'
                cotista_value = str(row_data['cotista']).strip().lower() if row_data['cotista'] else '0'
                cargo_nome = str(row_data['cargo']).strip() if row_data['cargo'] else ''
                nota = row_data['nota']

                pcd = 1 if pcd_value in ('1', 'true', 'sim') else 0
                cotista = 1 if cotista_value in ('1', 'true', 'sim') else 0
                
                if not nome or not numero_inscricao or ordem < 1 or not cargo_nome:
                    flash(f'Erro na linha {row_num}: Nome, número de inscrição, ordem válida e cargo são obrigatórios.', 'error')
                    row_num += 1
                    continue
                
                try:
                    nota = float(nota) if nota is not None else None
                    if nota is not None and nota < 0:
                        flash(f'Erro na linha {row_num}: A nota não pode ser negativa.', 'error')
                        row_num += 1
                        continue
                except (ValueError, TypeError):
                    flash(f'Erro na linha {row_num}: A nota deve ser um número válido.', 'error')
                    row_num += 1
                    continue
                
                # Normalizar cargo para comparação
                cargo_nome_normalized = cargo_nome.lower()
                if cargo_nome_normalized not in cargos_disponiveis:
                    flash(f'Erro na linha {row_num}: Cargo "{cargo_nome}" não encontrado no edital.', 'error')
                    row_num += 1
                    continue
                
                id_cargo = cargos_disponiveis[cargo_nome_normalized]
                
                candidato_existente = CandidatoController.obter_candidato(numero_inscricao)
                if candidato_existente:
                    flash(f'Erro na linha {row_num}: Número de inscrição {numero_inscricao} já cadastrado.', 'error')
                    row_num += 1
                    continue
                
                candidatos.append({
                    'nome': nome,
                    'numero_inscricao': numero_inscricao,
                    'id_cargo': id_cargo,
                    'ordem_nomeacao': ordem,
                    'pcd': pcd,
                    'cotista': cotista,
                    'id_edital': id_edital,
                    'nota': nota
                })
            except (ValueError, TypeError) as e:
                flash(f'Erro na linha {row_num}: Ordem deve ser um número inteiro, nota deve ser um número válido ou dados inválidos ({str(e)}).', 'error')
                row_num += 1
                continue
            row_num += 1

        if not candidatos:
            flash('Nenhum candidato válido encontrado no arquivo.', 'error')
            return redirect(url_for('edital.gerenciar_editais'))

        sucesso = CandidatoController.criar_candidatos_em_lote(candidatos)
        if sucesso:
            flash(f'{len(candidatos)} candidato(s) adicionado(s) com sucesso!', 'success')
        else:
            flash('Erro ao adicionar candidatos em lote.', 'error')

    except Exception as e:
        flash(f'Erro ao processar o arquivo: {str(e)}', 'error')

    return redirect(url_for('edital.gerenciar_editais'))